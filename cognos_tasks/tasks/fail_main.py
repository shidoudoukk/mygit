# -*- coding:utf-8 -*-

import sys
sys.path.append("..")
#from tasks import request
import os
import pandas as pd
import common.dataConnect as db
import datetime
import pandas as np
from xlutils.copy import copy
import xlrd,xlwt
from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.styles import Border, Alignment, numbers, Side, Font

outfile = 'D:/充值失败回访.xlsx'
bWin = True
winPath = 'E:\\Python\\demo\\config.ini'
linuxPath = '/ssd/script/config/timing_config/config.cnf'
if bWin:
    sConfigPath = winPath
else:
    sConfigPath = linuxPath

# 连接数据库
sqlCONN = db.clMySql(sConfigPath, 'default_glazedb', False)

start_date = (datetime.datetime.now() - datetime.timedelta(days = datetime.datetime.now().weekday()+7)).strftime('%Y-%m-%d')
end_date = (datetime.datetime.now() - datetime.timedelta(days = datetime.datetime.now().weekday())).strftime('%Y-%m-%d')
# start_date = '2019-04-01'
# end_date = '2019-04-08'
print('上周日期区间为%s-%s'%(start_date,end_date))

list_dir = os.listdir('D:\\output')
df_list = []
for i in list_dir:
    filePath = os.path.join("D:\\output", i)
    df = pd.read_excel(filePath, sheet_name=None)
    df = pd.concat(df.values(), sort=False)  # sort = False 不重新排序
    df_list.append(df)
df_excel = pd.concat(df_list, axis=0, ignore_index=True, sort=False)
mask = (df_excel['回访日期'] >= pd.to_datetime(start_date,format='%Y-%m-%d')) & (df_excel['回访日期'] < pd.to_datetime(end_date,format='%Y-%m-%d'))
df_excel = df_excel.loc[mask]

df_excel['回访日期+1'] = df_excel['回访日期'].apply(lambda x: (x-datetime.timedelta(days=-1)).strftime('%Y-%m-%d'))
df_excel['回访日期'] = df_excel['回访日期'].apply(lambda x: x.strftime('%Y-%m-%d'))

"""还是以user_id去重更为准确，不要用real_name为空，客服每次反馈格式不固定"""
#df_excel.dropna(axis=0, subset=["real_name"], inplace=True)
df_excel.drop_duplicates(subset='user_id', inplace=True)

sqlbase = """SELECT userId, DATE_FORMAT(rechargeTime,'%%Y-%%m-%%d') 充值成功日期, SUM(amount) 充值成功金额
FROM recharge_log
WHERE flag=1
AND TO_DAYS(rechargeTime) between TO_DAYS('%s') and TO_DAYS('%s')
GROUP BY userId, DATE_FORMAT(rechargeTime,'%%Y-%%m-%%d')
ORDER BY userId, DATE_FORMAT(rechargeTime,'%%Y-%%m-%%d')
"""%(start_date,end_date)

dfsql = pd.read_sql(sqlbase, sqlCONN.GetCONN())
dfmerge = pd.merge(df_excel,dfsql,how='left',
                    left_on=['user_id','回访日期'],
                    right_on=['userId','充值成功日期'],
                    suffixes=('_excel','_SQL'), indicator =True).drop(columns=['_merge','real_name', 'mobile','amount','recharge_log_id','remark','create_time','回访结果','userId'])

dfmerge = pd.merge(dfmerge, dfsql, how='left',left_on=['user_id', '回访日期+1'],
                   right_on=['userId', '充值成功日期'],suffixes=('_D0','_D1'), indicator =True).drop(columns=['_merge','userId', '充值成功日期_D0', '充值成功日期_D1', '回访日期+1'])

"""回访有效"""
recall_yuser = dfmerge[dfmerge['回访有效'] == '是']['user_id'].drop_duplicates().count()  #回访有效人数
D0sucessmem = dfmerge[dfmerge['回访有效'] == '是']['充值成功金额_D0'].count() #回访有效D0成功人数
D0sucessamo = dfmerge[dfmerge['回访有效'] == '是']['充值成功金额_D0'].sum() #回访有效D0充值金额

D1sucessmem = dfmerge[dfmerge['回访有效'] == '是']['充值成功金额_D1'].count() #回访有效D1成功人数
D1sucessamo = dfmerge[dfmerge['回访有效'] == '是']['充值成功金额_D1'].sum() #回访有效D1充值金额

D0sucessmemper = D0sucessmem/recall_yuser #回访有效D0成功人数占比
D1sucessmemper = D1sucessmem/recall_yuser #回访有效D1成功人数占比


"""回访无效"""
recall_nuser = dfmerge[dfmerge['回访有效'] != '是']['user_id'].drop_duplicates().count() #回访无效人数
D0failmem = dfmerge[dfmerge['回访有效'] != '是']['充值成功金额_D0'].count() #回访无效D0成功人数
D0failamo = dfmerge[dfmerge['回访有效'] != '是']['充值成功金额_D0'].sum() #回访无效D0充值金额

D1failmem = dfmerge[dfmerge['回访有效'] != '是']['充值成功金额_D1'].count() #回访无效D1成功人数
D1failamo = dfmerge[dfmerge['回访有效'] != '是']['充值成功金额_D1'].sum() #回访无效D1充值金额

D0failmemper = D0failmem/recall_nuser #回访无效D0成功人数占比
D1failmemper = D1failmem/recall_nuser #回访无效D1成功人数占比

# data = xlrd.open_workbook(outfile)
# table = data.sheets()[0]
# rows, cols = table.nrows, table.ncols
# workbook = copy(data)
# table_cal = workbook.get_sheet(0)
#
#
# for i in range(len(indata)):
#     print(i, indata[i])
#     table_cal.write(rows+1,i,str(indata[i]))

#workbook.save('D:/充值失败回访.xls')

wb = load_workbook(outfile)
ws=wb.worksheets[0]
nrows = ws.max_row
ncols = ws.max_column
indata = [nrows, recall_yuser, D0sucessmem, D0sucessamo, D1sucessmem, D1sucessamo, D0sucessmemper, D1sucessmemper,
          recall_nuser, D0failmem, D0failamo, D1failmem, D1failamo, D0failmemper, D1failmemper]
ws.append(indata)

border = Border(left=Side(border_style='thin',color='000000'),
right=Side(border_style='thin',color='000000'),
top=Side(border_style='thin',color='000000'),
bottom=Side(border_style='thin',color='000000'))

align = Alignment(horizontal='left',vertical='center',wrap_text=True)
font = Font(u'微软雅黑',size = 10,bold=False)

for i in range(ncols):
    ws.cell(row=nrows+1, column=i+1).border = border
    ws.cell(row=nrows+1, column=i+1).alignment = align
    ws.cell(row=nrows+1, column=i+1).font = font
per_index = [7,8,14,15]
for i in per_index:
    ws.cell(row=nrows+1, column=i).number_format='0.00%'
wb.save('D:/充值失败回访.xlsx')



