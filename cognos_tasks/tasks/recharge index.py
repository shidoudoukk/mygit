# -*- coding:utf-8 -*-

import sys
sys.path.append("..")
#from tasks import request
import os
import pandas as pd
import common.dataConnect as db
import datetime
import numpy as np
from xlutils.copy import copy
import xlrd,xlwt
from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.styles import Border, Alignment, numbers, Side, Font
import re

outpath = 'E:\\有利网\\充值指标\\'
bWin = True
winPath = 'E:\\Python\\demo\\config.ini'
linuxPath = '/ssd/script/config/timing_config/config.cnf'
if bWin:
    sConfigPath = winPath
else:
    sConfigPath = linuxPath

df_commit = pd.read_excel(outpath+'充值指标to产品.xlsx', sheet_name='快捷支付141')
max_date = df_commit['日期'].max()
max_date_str = datetime.datetime.strftime(max_date,'%Y%m%d')
rows = len(df_commit)+1
print('最大日期%s，行数%s'% (max_date_str, rows))


wb = load_workbook(outpath+'充值指标to产品.xlsx')
ws = wb['快捷支付141']
# nrows = ws.max_row
ncols = ws.max_column

"""表格设置"""
border = Border(left=Side(border_style='thin',color='BFBFBF'),
right=Side(border_style='thin',color='BFBFBF'),
top=Side(border_style='thin',color='BFBFBF'),
bottom=Side(border_style='thin',color='BFBFBF'))
align = Alignment(horizontal='center',vertical='center',wrap_text=True)
font = Font(u'微软雅黑',size = 9,bold=False)


# 获取充值指标数据
sqlCONN = db.clMySql(sConfigPath, 'default_glazedb', False)

list_dir = os.listdir(outpath)
df_list = []
i=1
for file in list_dir:
    filePath = os.path.join(outpath, file)
    if (not re.search("_1.xlsx", file) is None) and file[0:8]>max_date_str:
        df = pd.read_excel(filePath, sheet_name='datas1')
        df = df[df['平台编号']=='BJ020004'][['用户编号','支付金额']]
        pasauth_sucmem = df['用户编号'].drop_duplicates().count()
        passauth_sucnum = df['用户编号'].count()
        passauth_sucpri = df['支付金额'].sum()

        #print(rows+i)
        #df_commit.loc[rows+i, ('密码验证成功（人数）','密码验证成功（笔数）','密码验证成功（金额）')] = (pasauth_sucmem, passauth_sucnum, passauth_sucpri)
        #print('密码验证：',pasauth_sucmem,passauth_sucnum,passauth_sucpri)
        max_date = (max_date + datetime.timedelta(days=1))
        sqlbase = """SELECT DATE(rechargeTime) 日期,
        SUM(amount) 充值金额,
        '快捷支付' AS 支付方式,
        COUNT(DISTINCT userId) 下单成功人数,
        COUNT(rechargeLogId) 下单成功笔数,
        COUNT(DISTINCT CASE WHEN flag=1 THEN userId ELSE NULL END) 充值成功人数,
        COUNT(CASE WHEN flag=1 THEN rechargeLogId ELSE NULL END) 充值成功笔数,
        SUM(CASE WHEN flag=1 THEN amount ELSE NULL END) 充值成功金额,
        COUNT(CASE WHEN flag=0 THEN rechargeLogId ELSE NULL END) 充值失败笔数,
        SUM(CASE WHEN flag=0 THEN amount ELSE NULL END) 充值失败金额
        FROM recharge_log
        WHERE payType=141 AND TO_DAYS(rechargeTime) = TO_DAYS('%s')
        GROUP BY DATE(rechargeTime)
        """ % (max_date)
        df_recharge = pd.read_sql(sqlbase, sqlCONN.GetCONN())
        row = df_recharge.values[0]
        row = np.insert(row, 5, [pasauth_sucmem, passauth_sucnum, passauth_sucpri], 0)
        # df_commit.loc[rows+i, ('日期','充值金额','支付方式','下单成功（已经绑卡）-人数','下单成功（已经绑卡）-笔数',
        #                        '充值成功人数','充值成功（笔数）','充值成功金额','充值失败笔数','充值失败金额')] = df_recharge.values[0]

        ws.append(row.tolist())

        for j in range(ncols):
            ws.cell(row=rows + i, column=j + 1).border = border
            ws.cell(row=rows + i, column=j + 1).alignment = align
            ws.cell(row=rows + i, column=j + 1).font = font
        ws.cell(row=rows + i, column=1).number_format = "yyyy/m/d"

        i+=1
        #df_list.append(df)
# df_commit['有效提交率(人数）'] = df_commit['密码验证成功（人数）']/df_commit['下单成功（已经绑卡）-人数']
# df_commit['有效提交率(笔数）'] = df_commit['密码验证成功（笔数）']/df_commit['下单成功（已经绑卡）-笔数']
# df_commit['充值成功率(人数）'] = df_commit['充值成功人数']/df_commit['密码验证成功（人数）']
# df_commit['充值成功率(笔数）'] = df_commit['充值成功（笔数）']/df_commit['密码验证成功（笔数）']
# df_commit['充值成功/下单成功（人数）'] = df_commit['充值成功人数']/df_commit['下单成功（已经绑卡）-人数']
# df_commit['充值成功/下单成功（(笔数）'] = df_commit['充值成功（笔数）']/df_commit['下单成功（已经绑卡）-笔数']
# df_commit['有效提交率(金额）'] = df_commit['密码验证成功（金额）']/df_commit['充值金额']
# df_commit['充值成功率(金额）'] = df_commit['充值成功金额']/df_commit['密码验证成功（金额）']
# df_commit['充值成功/下单成功（(金额）'] = df_commit['充值成功金额']/df_commit['充值金额']
#
# per_column = ['有效提交率(人数）','有效提交率(笔数）','充值成功率(人数）','充值成功率(笔数）','充值成功/下单成功（人数）',
# '充值成功/下单成功（(笔数）','有效提交率(金额）','充值成功率(金额）','充值成功/下单成功（(金额）']
# for i in per_column:
#     df_commit[i] = df_commit[i].apply(lambda x: format(x, '.2%'))
# df_commit['日期'] = df_commit['日期'].apply(lambda x: x.strftime('%Y/%m/%d'))

wb.save(outpath+'充值指标to产品.xlsx')
print("done...")
#df_commit.to_excel(outpath+'test.xlsx',index=False)


